#!/usr/bin/env python3
"""
Generate expense report in standard Excel format
Compatible with Microsoft Excel, Google Sheets, and all spreadsheet applications
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def create_standard_excel_report():
    """Create a clean, professional Excel expense report"""

    # Expense data from receipts
    expenses = [
        {
            'date': '2019-11-20',
            'merchant': 'Harbor Lane Cafe',
            'description': 'Tacos Del Mar Shrimp, Especial Salad Chicken, Fountain Beverage',
            'category': 'Meals',
            'amount': 31.39,
            'currency': 'USD',
            'payment_type': 'Credit Card',
            'city': 'Chicago, IL',
            'receipt_file': 'ItemizedReceipt.jpg'
        },
        {
            'date': '2021-02-01',
            'merchant': 'Party Planner',
            'description': 'Event planning - Birthday invitation, Venue design and decorations, Catering',
            'category': 'Entertainment',
            'amount': 1920.00,
            'currency': 'USD',
            'payment_type': 'Credit Card',
            'city': 'Texas',
            'receipt_file': 'PartyInvoice.jpg'
        },
        {
            'date': '2024-01-01',
            'merchant': 'Nike Factory Store',
            'description': 'Nike Zoom Winflo, Air Force, Air Jordan 1 Mid BG',
            'category': 'Other',
            'amount': 144.97,
            'currency': 'USD',
            'payment_type': 'Credit Card',
            'city': 'Flushing Queens',
            'receipt_file': '348s.jpg'
        },
        {
            'date': '2012-12-10',
            'merchant': 'Albert Heijn',
            'description': 'Grocery items - AL Speculaas, Keukenrol, Surinam Rice, ES Witte Bol, Grillworst, Aardappelen',
            'category': 'Meals',
            'amount': 13.33,
            'currency': 'EUR',
            'payment_type': 'Credit Card',
            'city': 'Netherlands',
            'receipt_file': 'VoorbeeldAH.jpeg'
        }
    ]

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Add title
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "EXPENSE REPORT"
    title_cell.font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells('A1:I1')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Add report date
    report_date_cell = ws.cell(row=2, column=1)
    report_date_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    report_date_cell.font = Font(italic=True, size=10)
    ws.merge_cells('A2:I2')
    report_date_cell.alignment = Alignment(horizontal='center')

    # Blank row
    ws.row_dimensions[3].height = 10

    # Define headers
    headers = [
        "Date",
        "Merchant",
        "Description",
        "Category",
        "Amount",
        "Currency",
        "Payment",
        "City",
        "Receipt"
    ]

    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Write headers (row 4)
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[header_row].height = 25

    # Alternating row colors
    light_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Write expense data starting from row 5
    start_row = 5
    for idx, expense in enumerate(expenses):
        row = start_row + idx

        # Alternate row colors
        row_fill = light_fill if idx % 2 == 0 else white_fill

        # Date
        date_cell = ws.cell(row=row, column=1, value=expense['date'])
        date_cell.alignment = Alignment(horizontal='center')
        date_cell.border = thin_border
        date_cell.fill = row_fill

        # Merchant
        merchant_cell = ws.cell(row=row, column=2, value=expense['merchant'])
        merchant_cell.border = thin_border
        merchant_cell.fill = row_fill

        # Description
        desc_cell = ws.cell(row=row, column=3, value=expense['description'])
        desc_cell.alignment = Alignment(wrap_text=True)
        desc_cell.border = thin_border
        desc_cell.fill = row_fill

        # Category
        cat_cell = ws.cell(row=row, column=4, value=expense['category'])
        cat_cell.border = thin_border
        cat_cell.fill = row_fill

        # Amount
        amount_cell = ws.cell(row=row, column=5, value=float(expense['amount']))
        amount_cell.number_format = '#,##0.00'
        amount_cell.alignment = Alignment(horizontal='right')
        amount_cell.border = thin_border
        amount_cell.fill = row_fill

        # Currency
        curr_cell = ws.cell(row=row, column=6, value=expense['currency'])
        curr_cell.alignment = Alignment(horizontal='center')
        curr_cell.border = thin_border
        curr_cell.fill = row_fill

        # Payment Type
        pay_cell = ws.cell(row=row, column=7, value=expense['payment_type'])
        pay_cell.border = thin_border
        pay_cell.fill = row_fill

        # City
        city_cell = ws.cell(row=row, column=8, value=expense['city'])
        city_cell.border = thin_border
        city_cell.fill = row_fill

        # Receipt File
        receipt_cell = ws.cell(row=row, column=9, value=expense['receipt_file'])
        receipt_cell.border = thin_border
        receipt_cell.fill = row_fill

    # Add summary section
    summary_row = start_row + len(expenses) + 2

    # Summary header
    summary_header = ws.cell(row=summary_row, column=1)
    summary_header.value = "SUMMARY"
    summary_header.font = Font(bold=True, size=12, color="1F4E78")
    ws.merge_cells(f'A{summary_row}:I{summary_row}')
    summary_header.alignment = Alignment(horizontal='left')

    summary_row += 1

    # Calculate totals
    usd_total = sum(e['amount'] for e in expenses if e['currency'] == 'USD')
    eur_total = sum(e['amount'] for e in expenses if e['currency'] == 'EUR')

    # USD Total
    usd_label = ws.cell(row=summary_row, column=4)
    usd_label.value = "Total USD:"
    usd_label.font = Font(bold=True)
    usd_label.alignment = Alignment(horizontal='right')

    usd_amount = ws.cell(row=summary_row, column=5)
    usd_amount.value = usd_total
    usd_amount.font = Font(bold=True, size=11)
    usd_amount.number_format = '$#,##0.00'
    usd_amount.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    usd_amount.border = Border(
        top=Side(style='double', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    ws.cell(row=summary_row, column=6, value='USD').font = Font(bold=True)

    # EUR Total
    summary_row += 1
    eur_label = ws.cell(row=summary_row, column=4)
    eur_label.value = "Total EUR:"
    eur_label.font = Font(bold=True)
    eur_label.alignment = Alignment(horizontal='right')

    eur_amount = ws.cell(row=summary_row, column=5)
    eur_amount.value = eur_total
    eur_amount.font = Font(bold=True, size=11)
    eur_amount.number_format = '€#,##0.00'
    eur_amount.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    eur_amount.border = Border(
        top=Side(style='double', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    ws.cell(row=summary_row, column=6, value='EUR').font = Font(bold=True)

    # Category breakdown
    summary_row += 2
    category_header = ws.cell(row=summary_row, column=1)
    category_header.value = "BY CATEGORY"
    category_header.font = Font(bold=True, size=11, color="1F4E78")

    summary_row += 1

    # Calculate by category
    by_category = {}
    for exp in expenses:
        cat = exp['category']
        if cat not in by_category:
            by_category[cat] = {'count': 0, 'total': 0}
        by_category[cat]['count'] += 1
        if exp['currency'] == 'USD':
            by_category[cat]['total'] += exp['amount']

    for category, data in sorted(by_category.items()):
        cat_label = ws.cell(row=summary_row, column=2)
        cat_label.value = category
        cat_label.font = Font(size=10)

        cat_count = ws.cell(row=summary_row, column=3)
        cat_count.value = f"{data['count']} expense(s)"
        cat_count.font = Font(size=10)

        if data['total'] > 0:
            cat_amount = ws.cell(row=summary_row, column=5)
            cat_amount.value = data['total']
            cat_amount.number_format = '$#,##0.00'
            cat_amount.font = Font(size=10)

        summary_row += 1

    # Adjust column widths for optimal viewing
    column_widths = {
        'A': 12,   # Date
        'B': 22,   # Merchant
        'C': 45,   # Description
        'D': 15,   # Category
        'E': 12,   # Amount
        'F': 10,   # Currency
        'G': 14,   # Payment
        'H': 18,   # City
        'I': 22    # Receipt
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Set row heights for data rows
    for row in range(start_row, start_row + len(expenses)):
        ws.row_dimensions[row].height = 35

    # Freeze panes (freeze header row)
    ws.freeze_panes = 'A5'

    # Save workbook
    output_file = 'expense_report_standard.xlsx'
    wb.save(output_file)

    print("\n" + "=" * 70)
    print("Standard Excel Expense Report Generated!")
    print("=" * 70)
    print(f"\nFile: {output_file}")
    print(f"Format: Microsoft Excel (.xlsx)")
    print(f"Expenses: {len(expenses)}")
    print(f"\nTotals:")
    print(f"  USD: ${usd_total:,.2f}")
    print(f"  EUR: €{eur_total:,.2f}")
    print(f"\n✓ Ready to open in Microsoft Excel, Google Sheets, or any spreadsheet app")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    create_standard_excel_report()
