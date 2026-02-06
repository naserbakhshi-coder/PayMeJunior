#!/usr/bin/env python3
"""
PayMeJunior Agent - Expense Report Generator
Processes receipt images and creates SAP Concur-ready Excel expense reports
"""

import os
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import anthropic
import base64
import json
from typing import Dict, List, Optional


class PayMeJuniorAgent:
    """Agent responsible for creating expense reports from receipts"""

    def __init__(self, receipts_folder: str = "."):
        self.receipts_folder = Path(receipts_folder)
        self.client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
        self.expenses = []

    def get_receipt_files(self) -> List[Path]:
        """Get all receipt image files from the folder"""
        image_extensions = {'.jpg', '.jpeg', '.png', '.pdf', '.gif', '.bmp'}
        receipt_files = []

        for file in self.receipts_folder.iterdir():
            if file.is_file() and file.suffix.lower() in image_extensions:
                receipt_files.append(file)

        return sorted(receipt_files)

    def encode_image(self, image_path: Path) -> tuple[str, str]:
        """Encode image to base64 and determine media type"""
        with open(image_path, 'rb') as f:
            image_data = base64.standard_b64encode(f.read()).decode('utf-8')

        # Determine media type
        ext = image_path.suffix.lower()
        media_types = {
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
            '.bmp': 'image/bmp'
        }
        media_type = media_types.get(ext, 'image/jpeg')

        return image_data, media_type

    def extract_expense_data(self, receipt_path: Path) -> Optional[Dict]:
        """Extract expense information from a receipt using Claude AI"""
        print(f"Processing {receipt_path.name}...")

        try:
            image_data, media_type = self.encode_image(receipt_path)

            response = self.client.messages.create(
                model="claude-opus-4-5-20251101",
                max_tokens=1024,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": media_type,
                                    "data": image_data
                                }
                            },
                            {
                                "type": "text",
                                "text": """Analyze this receipt and extract the following information in JSON format:
{
  "date": "YYYY-MM-DD format",
  "merchant": "Merchant/Vendor name",
  "description": "Brief description of the expense",
  "amount": "Total amount as decimal number only",
  "currency": "Currency code (USD, EUR, etc.)",
  "category": "Expense category (Meals, Transportation, Office Supplies, Entertainment, Lodging, Other)",
  "payment_type": "Credit Card",
  "city": "City if available",
  "items": "Brief list of items purchased"
}

Please extract the exact values from the receipt. For the category, choose the most appropriate one based on what was purchased. Return ONLY the JSON object, no other text."""
                            }
                        ]
                    }
                ]
            )

            # Extract JSON from response
            content = response.content[0].text.strip()
            # Remove markdown code blocks if present
            if content.startswith('```'):
                content = content.split('```')[1]
                if content.startswith('json'):
                    content = content[4:]
                content = content.strip()

            expense_data = json.loads(content)
            expense_data['receipt_file'] = receipt_path.name

            print(f"  ✓ Extracted: {expense_data['merchant']} - {expense_data['currency']} {expense_data['amount']}")

            return expense_data

        except Exception as e:
            print(f"  ✗ Error processing {receipt_path.name}: {str(e)}")
            return None

    def process_all_receipts(self):
        """Process all receipts in the folder"""
        receipt_files = self.get_receipt_files()

        if not receipt_files:
            print("No receipt files found in the folder.")
            return

        print(f"\nFound {len(receipt_files)} receipt(s) to process\n")
        print("-" * 60)

        for receipt_file in receipt_files:
            expense_data = self.extract_expense_data(receipt_file)
            if expense_data:
                self.expenses.append(expense_data)

        print("-" * 60)
        print(f"\nSuccessfully processed {len(self.expenses)} out of {len(receipt_files)} receipts\n")

    def create_concur_excel(self, output_file: str = "expense_report.xlsx"):
        """Create SAP Concur-ready Excel expense report"""

        if not self.expenses:
            print("No expenses to export.")
            return

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        # Define headers based on SAP Concur format
        headers = [
            "Expense Date",
            "Merchant/Vendor",
            "Description",
            "Expense Type",
            "Amount",
            "Currency",
            "Payment Type",
            "City",
            "Receipt File"
        ]

        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Write expense data
        for row, expense in enumerate(self.expenses, start=2):
            ws.cell(row=row, column=1, value=expense.get('date', '')).border = border
            ws.cell(row=row, column=2, value=expense.get('merchant', '')).border = border
            ws.cell(row=row, column=3, value=expense.get('description', '')).border = border
            ws.cell(row=row, column=4, value=expense.get('category', 'Other')).border = border

            # Format amount as number
            amount_cell = ws.cell(row=row, column=5, value=float(expense.get('amount', 0)))
            amount_cell.number_format = '#,##0.00'
            amount_cell.border = border

            ws.cell(row=row, column=6, value=expense.get('currency', 'USD')).border = border
            ws.cell(row=row, column=7, value=expense.get('payment_type', 'Credit Card')).border = border
            ws.cell(row=row, column=8, value=expense.get('city', '')).border = border
            ws.cell(row=row, column=9, value=expense.get('receipt_file', '')).border = border

        # Add total row
        total_row = len(self.expenses) + 2
        total_cell = ws.cell(row=total_row, column=4)
        total_cell.value = "TOTAL:"
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal='right')

        # Calculate total (note: this is simplified - in reality you'd need currency conversion)
        total_amount_cell = ws.cell(row=total_row, column=5)
        total_amount_cell.value = f"=SUM(E2:E{total_row-1})"
        total_amount_cell.font = Font(bold=True)
        total_amount_cell.number_format = '#,##0.00'
        total_amount_cell.border = Border(top=Side(style='double'))

        # Adjust column widths
        column_widths = {
            'A': 15,  # Date
            'B': 25,  # Merchant
            'C': 35,  # Description
            'D': 20,  # Expense Type
            'E': 12,  # Amount
            'F': 10,  # Currency
            'G': 15,  # Payment Type
            'H': 15,  # City
            'I': 25   # Receipt File
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Save workbook
        output_path = self.receipts_folder / output_file
        wb.save(output_path)

        print(f"✓ Expense report created: {output_path}")
        print(f"  Total expenses: {len(self.expenses)}")
        print(f"  Ready for SAP Concur upload!")

        return str(output_path)

    def generate_summary(self):
        """Generate a summary of expenses"""
        if not self.expenses:
            return

        print("\n" + "=" * 60)
        print("EXPENSE REPORT SUMMARY")
        print("=" * 60)

        # Group by category
        by_category = {}
        by_currency = {}

        for expense in self.expenses:
            category = expense.get('category', 'Other')
            currency = expense.get('currency', 'USD')
            amount = float(expense.get('amount', 0))

            by_category[category] = by_category.get(category, 0) + amount
            by_currency[currency] = by_currency.get(currency, 0) + amount

        print("\nBy Category:")
        for category, total in sorted(by_category.items()):
            print(f"  {category:.<30} ${total:>10.2f}")

        print("\nBy Currency:")
        for currency, total in sorted(by_currency.items()):
            print(f"  {currency:.<30} {total:>10.2f}")

        print("\n" + "=" * 60 + "\n")


def main():
    """Main execution function"""
    print("\n" + "=" * 60)
    print("PayMeJunior Agent - Expense Report Generator")
    print("=" * 60 + "\n")

    # Check for API key
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ERROR: ANTHROPIC_API_KEY environment variable not set.")
        print("Please set your API key: export ANTHROPIC_API_KEY='your-key-here'")
        sys.exit(1)

    # Get receipts folder from command line or use current directory
    receipts_folder = sys.argv[1] if len(sys.argv) > 1 else "."

    # Create agent
    agent = PayMeJuniorAgent(receipts_folder)

    # Process receipts
    agent.process_all_receipts()

    # Create Excel report
    if agent.expenses:
        agent.create_concur_excel()
        agent.generate_summary()
    else:
        print("No expenses were successfully processed. No report generated.")


if __name__ == "__main__":
    main()
