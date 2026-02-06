"""
Excel report generator for SAP Concur format
Refactored from original paymejunior_agent.py
"""
import io
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelGenerator:
    """
    Generates SAP Concur-compatible Excel expense reports.
    """

    # SAP Concur column headers
    HEADERS = [
        "Expense Date",
        "Merchant/Vendor",
        "Description",
        "Expense Type",
        "Amount",
        "Currency",
        "Payment Type",
        "City",
        "Receipt File"
    ]

    # Column widths for readability
    COLUMN_WIDTHS = {
        'A': 15,   # Date
        'B': 25,   # Merchant
        'C': 35,   # Description
        'D': 20,   # Expense Type
        'E': 12,   # Amount
        'F': 10,   # Currency
        'G': 15,   # Payment Type
        'H': 15,   # City
        'I': 25    # Receipt File
    }

    def __init__(self):
        # Style definitions
        self.header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_report(
        self,
        expenses: list[dict],
        report_name: str = "expense_report"
    ) -> bytes:
        """
        Generate SAP Concur-ready Excel expense report.

        Args:
            expenses: List of expense dictionaries
            report_name: Name for the report (used in sheet title)

        Returns:
            Excel file as bytes
        """
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        # Write headers
        self._write_headers(ws)

        # Write expense data
        for row, expense in enumerate(expenses, start=2):
            self._write_expense_row(ws, row, expense)

        # Add total row
        if expenses:
            self._write_total_row(ws, len(expenses))

        # Adjust column widths
        self._set_column_widths(ws)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _write_headers(self, ws) -> None:
        """Write styled headers to worksheet"""
        for col, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

    def _write_expense_row(self, ws, row: int, expense: dict) -> None:
        """Write a single expense row to worksheet"""
        # Map expense fields to columns
        ws.cell(row=row, column=1, value=expense.get('date', '')).border = self.border
        ws.cell(row=row, column=2, value=expense.get('merchant', '')).border = self.border
        ws.cell(row=row, column=3, value=expense.get('description', '')).border = self.border
        ws.cell(row=row, column=4, value=expense.get('category', 'Other')).border = self.border

        # Format amount as number
        amount_cell = ws.cell(row=row, column=5)
        amount_value = expense.get('amount', 0)
        if isinstance(amount_value, str):
            amount_value = float(amount_value.replace('$', '').replace(',', ''))
        amount_cell.value = float(amount_value)
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = self.border

        ws.cell(row=row, column=6, value=expense.get('currency', 'USD')).border = self.border
        ws.cell(row=row, column=7, value=expense.get('payment_type', 'Credit Card')).border = self.border
        ws.cell(row=row, column=8, value=expense.get('city', '')).border = self.border
        ws.cell(row=row, column=9, value=expense.get('receipt_path', '')).border = self.border

    def _write_total_row(self, ws, expense_count: int) -> None:
        """Write total row with SUM formula"""
        total_row = expense_count + 2

        # "TOTAL:" label
        total_label = ws.cell(row=total_row, column=4)
        total_label.value = "TOTAL:"
        total_label.font = Font(bold=True)
        total_label.alignment = Alignment(horizontal='right')

        # Sum formula
        total_amount = ws.cell(row=total_row, column=5)
        total_amount.value = f"=SUM(E2:E{total_row - 1})"
        total_amount.font = Font(bold=True)
        total_amount.number_format = '#,##0.00'
        total_amount.border = Border(top=Side(style='double'))

    def _set_column_widths(self, ws) -> None:
        """Set column widths for readability"""
        for col, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width

    def generate_summary(self, expenses: list[dict]) -> dict:
        """
        Generate expense summary by category and currency.

        Args:
            expenses: List of expense dictionaries

        Returns:
            Summary dict with by_category, by_currency, and grand_totals
        """
        by_category = {}
        by_currency = {}

        for expense in expenses:
            category = expense.get('category', 'Other')
            currency = expense.get('currency', 'USD')
            amount = float(expense.get('amount', 0))

            # Group by category
            if category not in by_category:
                by_category[category] = {"count": 0, "total": 0}
            by_category[category]["count"] += 1
            by_category[category]["total"] += amount

            # Group by currency
            if currency not in by_currency:
                by_currency[currency] = {"count": 0, "total": 0}
            by_currency[currency]["count"] += 1
            by_currency[currency]["total"] += amount

        return {
            "by_category": by_category,
            "by_currency": by_currency,
            "grand_totals": {
                "total_expenses": len(expenses),
                "currencies": list(by_currency.keys())
            }
        }


# Singleton instance
_excel_generator: Optional[ExcelGenerator] = None


def get_excel_generator() -> ExcelGenerator:
    """Get or create the Excel generator singleton"""
    global _excel_generator
    if _excel_generator is None:
        _excel_generator = ExcelGenerator()
    return _excel_generator
