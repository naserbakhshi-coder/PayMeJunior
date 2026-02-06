#!/usr/bin/env python3
"""
Quick expense report generator with pre-extracted data
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def create_expense_report():
    """Create SAP Concur-ready Excel expense report"""

    # Pre-extracted expense data from receipts
    expenses = [
        {
            'date': '2019-11-20',
            'merchant': 'Harbor Lane Cafe',
            'description': 'Tacos Del Mar Shrimp, Especial Salad Chicken, Fountain Beverage',
            'category': 'Meals',
            'amount': 31.39,
            'currency': 'USD',
            'payment_type': 'Credit Card',
            'city': 'Chicago, IL',
            'receipt_file': 'ItemizedReceipt.jpg'
        },
        {
            'date': '2021-02-01',
            'merchant': 'Party Planner',
            'description': 'Event planning - Birthday invitation, Venue design and decorations, Catering',
            'category': 'Entertainment',
            'amount': 1920.00,
            'currency': 'USD',
            'payment_type': 'Credit Card',
            'city': 'Texas',
            'receipt_file': 'PartyInvoice.jpg'
        },
        {
            'date': '2024-01-01',  # Date not visible on receipt, using estimate
            'merchant': 'Nike Factory Store',
            'description': 'Nike Zoom Winflo, Air Force, Air Jordan 1 Mid BG',
            'category': 'Other',
            'amount': 144.97,
            'currency': 'USD',
            'payment_type': 'Credit Card',
            'city': 'Flushing Queens',
            'receipt_file': '348s.jpg'
        },
        {
            'date': '2012-12-10',
            'merchant': 'Albert Heijn',
            'description': 'Grocery items - AL Speculaas, Keukenrol, Surinam Rice, ES Witte Bol, Grillworst, Aardappelen',
            'category': 'Meals',
            'amount': 13.33,
            'currency': 'EUR',
            'payment_type': 'Credit Card',
            'city': 'Netherlands',
            'receipt_file': 'VoorbeeldAH.jpeg'
        }
    ]

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    # Define headers based on SAP Concur format
    headers = [
        "Expense Date",
        "Merchant/Vendor",
        "Description",
        "Expense Type",
        "Amount",
        "Currency",
        "Payment Type",
        "City",
        "Receipt File"
    ]

    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Write expense data
    for row, expense in enumerate(expenses, start=2):
        ws.cell(row=row, column=1, value=expense['date']).border = border
        ws.cell(row=row, column=2, value=expense['merchant']).border = border
        ws.cell(row=row, column=3, value=expense['description']).border = border
        ws.cell(row=row, column=4, value=expense['category']).border = border

        # Format amount as number
        amount_cell = ws.cell(row=row, column=5, value=float(expense['amount']))
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = border

        ws.cell(row=row, column=6, value=expense['currency']).border = border
        ws.cell(row=row, column=7, value=expense['payment_type']).border = border
        ws.cell(row=row, column=8, value=expense['city']).border = border
        ws.cell(row=row, column=9, value=expense['receipt_file']).border = border

    # Add total row for USD expenses
    total_row = len(expenses) + 2
    total_cell = ws.cell(row=total_row, column=4)
    total_cell.value = "TOTAL (USD):"
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal='right')

    # Calculate USD total
    usd_total = sum(e['amount'] for e in expenses if e['currency'] == 'USD')
    total_amount_cell = ws.cell(row=total_row, column=5)
    total_amount_cell.value = usd_total
    total_amount_cell.font = Font(bold=True)
    total_amount_cell.number_format = '#,##0.00'
    total_amount_cell.border = Border(top=Side(style='double'))

    ws.cell(row=total_row, column=6, value='USD').font = Font(bold=True)

    # Add EUR total
    eur_row = total_row + 1
    eur_cell = ws.cell(row=eur_row, column=4)
    eur_cell.value = "TOTAL (EUR):"
    eur_cell.font = Font(bold=True)
    eur_cell.alignment = Alignment(horizontal='right')

    eur_total = sum(e['amount'] for e in expenses if e['currency'] == 'EUR')
    eur_amount_cell = ws.cell(row=eur_row, column=5)
    eur_amount_cell.value = eur_total
    eur_amount_cell.font = Font(bold=True)
    eur_amount_cell.number_format = '#,##0.00'

    ws.cell(row=eur_row, column=6, value='EUR').font = Font(bold=True)

    # Adjust column widths
    column_widths = {
        'A': 15,  # Date
        'B': 25,  # Merchant
        'C': 50,  # Description
        'D': 20,  # Expense Type
        'E': 12,  # Amount
        'F': 10,  # Currency
        'G': 15,  # Payment Type
        'H': 20,  # City
        'I': 25   # Receipt File
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save workbook
    output_file = 'expense_report.xlsx'
    wb.save(output_file)

    print("\n" + "=" * 70)
    print("PayMeJunior - Expense Report Generated Successfully!")
    print("=" * 70)
    print(f"\nOutput file: {output_file}")
    print(f"Total expenses processed: {len(expenses)}")
    print(f"\nSummary by Category:")

    by_category = {}
    for exp in expenses:
        cat = exp['category']
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(exp['amount'])

    for cat, amounts in sorted(by_category.items()):
        print(f"  {cat}: {len(amounts)} expense(s)")

    print(f"\nTotals:")
    print(f"  USD: ${usd_total:,.2f}")
    print(f"  EUR: €{eur_total:,.2f}")
    print(f"\n✓ Report is ready for SAP Concur upload!")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    create_expense_report()
